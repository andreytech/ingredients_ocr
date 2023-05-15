import pandas as pd
import os
import os.path as _path
import openpyxl
import threading 

class Excel_table:
    def __init__(self, path : str, sheet_name):
        self.path = path
        self.focused_sheet = sheet_name
        
        self.locker = threading.Lock()
        #creating
        if not _path.exists(path): 
            wb = openpyxl.Workbook()
            wb.active.title = self.focused_sheet
            wb.save(self.path)
        
            self.dataframe = pd.DataFrame( columns=["Название", "Ссылка", "Бренд", "Описание", "Ссылки на фото"])
            self.dataframe.to_excel(self.path, sheet_name= self.focused_sheet, index=False)

            self.success = True
        else :
            workbook = openpyxl.load_workbook(path)
            
            if self.focused_sheet not in workbook.sheetnames:
                
                self.dataframe = pd.DataFrame( columns=["Название", "Ссылка", "Бренд", "Описание", "Ссылки на фото"])
                self.dataframe.to_excel(self.path, sheet_name= self.focused_sheet, index=False)
                
            else :
                self.dataframe = pd.read_excel(path, self.focused_sheet)
        print(self.dataframe)

    def add_product(self, params):
        print(params)
        params_ = {
            "Название" : params["name"], 
            "Ссылка" : params["url"], 
            "Бренд" : params["brend"], 
            "Описание" : escape_xlsx_string( params["description"]), 
            "Ссылки на фото" : params["photo-urls"]
        }
        self.dataframe.loc[len(self.dataframe)] = params_


    def save_table(self):
        try:
            print(self.dataframe)
            self.dataframe.to_excel(self.path, sheet_name = self.focused_sheet, index= False)
        except Exception as err:
            print (err)


def escape_xlsx_char(ch):
	illegal_xlsx_chars = {
	'\x00':'\\x00',	#	NULL
	'\x01':'\\x01',	#	SOH
	'\x02':'\\x02',	#	STX
	'\x03':'\\x03',	#	ETX
	'\x04':'\\x04',	#	EOT
	'\x05':'\\x05',	#	ENQ
	'\x06':'\\x06',	#	ACK
	'\x07':'\\x07',	#	BELL
	'\x08':'\\x08',	#	BS
	'\x0b':'\\x0b',	#	VT
	'\x0c':'\\x0c',	#	FF
	'\x0e':'\\x0e',	#	SO
	'\x0f':'\\x0f',	#	SI
	'\x10':'\\x10',	#	DLE
	'\x11':'\\x11',	#	DC1
	'\x12':'\\x12',	#	DC2
	'\x13':'\\x13',	#	DC3
	'\x14':'\\x14',	#	DC4
	'\x15':'\\x15',	#	NAK
	'\x16':'\\x16',	#	SYN
	'\x17':'\\x17',	#	ETB
	'\x18':'\\x18',	#	CAN
	'\x19':'\\x19',	#	EM
	'\x1a':'\\x1a',	#	SUB
	'\x1b':'\\x1b',	#	ESC
	'\x1c':'\\x1c',	#	FS
	'\x1d':'\\x1d',	#	GS
	'\x1e':'\\x1e',	#	RS
	'\x1f':'\\x1f'}	#	US
	
	if ch in illegal_xlsx_chars:
		return illegal_xlsx_chars[ch]
	
	return ch
	
#
# Wraps the function escape_xlsx_char(ch).
def escape_xlsx_string(string):
	
	return ''.join([escape_xlsx_char(ch) for ch in string])